# -*- coding: utf-8 -*-
"""
Export tennis scheduler results to Excel, using template format.
Based on schedule_20260625.xlsx template.
Date/time are user-configurable, player names get unique colors.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import json, sys, os, copy
from datetime import datetime

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "template.xlsx")

PLAYER_COLORS = [
    "ECDAD4", "E6ECD4", "D1E0EF", "EED2EE",
    "E0EFD1", "D2EEEE", "F1F1CF", "E0D1EF",
    "FCE4D6", "D6E4FC", "E2EFDA", "DAEEF3",
    "F2DCDB", "DBEEDB", "E4DFEC", "FDE9D9",
    "D9E2F3", "E2F0D9", "EEE7D2", "D1EFE0",
    "EFD1E0", "D2EED2", "D9EED2", "D2D2EE",
    "EED2D2", "EFE0D1"
]

def parse_time(t_str):
    """Parse '19:00' -> (19, 0)"""
    parts = t_str.split(':')
    return int(parts[0]), int(parts[1]) if len(parts) > 1 else 0

def format_time(h, m):
    return f"{h:02d}:{m:02d}"

def generate_time_slots(venue_time, duration_min, num_slots):
    """Generate per-round time slots from venue time range."""
    parts = venue_time.split('-')
    start_str = parts[0].strip()
    end_str = parts[1].strip() if len(parts) > 1 else parts[0].strip()
    sh, sm = parse_time(start_str)
    start_total = sh * 60 + sm
    
    slots = []
    for i in range(num_slots):
        s = start_total + i * duration_min
        e = s + duration_min
        sh_i, sm_i = s // 60, s % 60
        eh_i, em_i = e // 60, e % 60
        slots.append(f"{format_time(sh_i, sm_i)}-{format_time(eh_i, em_i)}")
    return slots

def export(matches, output_path, date_str="6月25日", venue_time="19:00-22:00",
           duration=30, male_names=None, female_names=None):
    """
    Export matches to Excel using template format.
    
    matches: list of [round, type, team1, team2, ...]
    date_str: e.g. "6月25日"
    venue_time: e.g. "19:00-22:00"
    duration: minutes per round
    """
    # Load template if available, otherwise create from scratch
    if os.path.exists(TEMPLATE_PATH):
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
        ws = wb['对局安排']
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "对局安排"
    
    # Helper to get player name
    def get_name(match, team_idx, player_idx):
        r, mtype, t1, t2 = match[:4]
        team = t1 if team_idx == 0 else t2
        if mtype in ('DB', 'WB'):
            idx = team[player_idx]
            gender = 'M' if mtype == 'DB' else 'W'
            names = male_names if gender == 'M' else female_names
            return names[idx] if names and idx < len(names) else f"{gender}{idx+1}"
        elif mtype == 'XW':
            # MX side: (man, woman_mx), WB side: (woman0, woman1)
            if team_idx == 0:
                idx = team[player_idx]
                return male_names[idx] if male_names and idx < len(male_names) else f"M{idx+1}"
            else:
                idx = team[player_idx]
                return female_names[idx] if female_names and idx < len(female_names) else f"W{idx+1}"
        elif mtype == 'XD':
            # MX side: (man_mx, woman), DB side: (man0, man1)
            if team_idx == 0:
                idx = team[player_idx]
                return male_names[idx] if male_names and idx < len(male_names) else f"M{idx+1}"
            else:
                idx = team[player_idx]
                return male_names[idx] if male_names and idx < len(male_names) else f"M{idx+1}"
        else:  # MX
            if team_idx == 0:
                idx = team[player_idx]
                return male_names[idx] if male_names and idx < len(male_names) else f"M{idx+1}"
            else:
                idx = team[player_idx]
                return female_names[idx] if female_names and idx < len(female_names) else f"W{idx+1}"
    
    # Assign colors to unique names
    name_colors = {}
    for m in matches:
        for ti in range(2):
            for pi in range(2):
                name = get_name(m, ti, pi)
                if name not in name_colors:
                    cidx = len(name_colors) % len(PLAYER_COLORS)
                    name_colors[name] = PLAYER_COLORS[cidx]
    
    # Group matches by round
    rounds = {}
    for m in matches:
        r = m[0]
        if r not in rounds: rounds[r] = []
        rounds[r].append(m)
    
    sorted_rounds = sorted(rounds.keys())
    num_rounds = len(sorted_rounds)
    max_courts = max(len(v) for v in rounds.values())
    
    # Generate time slots
    time_slots = generate_time_slots(venue_time, duration, num_rounds)
    
    # ---- Row 1: Title ----
    ws.merge_cells('A1:N1')
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f" {date_str} 场地时间：{venue_time}"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    
    # ---- Row 2: Headers (keep template style) ----
    default_hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    default_hdr_font = Font(color="FFFFFF", bold=True, size=11)
    hdr_align = Alignment(horizontal="center", vertical="center")
    
    headers = ['时间段', '场次', '场地1', None, None, '场地2', None, None, 
               '场地3', None, None, '场地4', None, None]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=ci)
        cell.value = h
        cell.fill = default_hdr_fill
        cell.font = default_hdr_font
        cell.alignment = hdr_align
    
    # Merge header cells
    if not os.path.exists(TEMPLATE_PATH):
        ws.merge_cells('C2:E2')
        ws.merge_cells('F2:H2')
        ws.merge_cells('I2:K2')
        ws.merge_cells('L2:N2')
    ws.row_dimensions[2].height = 25
    
    # ---- Set column widths ----
    col_widths = {1: 14, 2: 8, 3: 8, 4: 18, 5: 18, 6: 8, 7: 18, 8: 18,
                  9: 8, 10: 18, 11: 18, 12: 8, 13: 18, 14: 18}
    for c, w in col_widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    
    # ---- Data rows (start from row 3) ----
    type_label = {'DB': '男双', 'WB': '女双', 'MX': '混双', 'XW': 'MX×WB', 'XD': 'MX×DB'}
    thin_border = None  # Keep template borders
    
    current_row = 3
    
    for round_idx, rnd in enumerate(sorted_rounds):
        ms = rounds[rnd]
        data_row1 = current_row
        data_row2 = current_row + 1
        
        # Time slot (merged 2 rows)
        ws.merge_cells(start_row=data_row1, start_column=1, end_row=data_row2, end_column=1)
        time_cell = ws.cell(row=data_row1, column=1)
        time_cell.value = time_slots[round_idx] if round_idx < len(time_slots) else ""
        time_cell.font = Font(size=11)
        time_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Round number (merged 2 rows)
        ws.merge_cells(start_row=data_row1, start_column=2, end_row=data_row2, end_column=2)
        rnd_cell = ws.cell(row=data_row1, column=2)
        rnd_cell.value = rnd
        rnd_cell.font = Font(size=11)
        rnd_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Each court
        for ci in range(max_courts):
            col_start = 3 + ci * 3
            
            if ci < len(ms):
                m = ms[ci]
                t_label = type_label.get(m[1], m[1])
                
                # Type (merged 2 rows)
                ws.merge_cells(start_row=data_row1, start_column=col_start,
                              end_row=data_row2, end_column=col_start)
                type_cell = ws.cell(row=data_row1, column=col_start)
                type_cell.value = t_label
                type_cell.font = Font(size=11)
                type_cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Team 1 players (row 1)
                for pi in range(2):
                    name = get_name(m, 0, pi)
                    cell = ws.cell(row=data_row1, column=col_start + 1 + pi)
                    cell.value = name
                    cell.font = Font(bold=True, size=10)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if name in name_colors:
                        cell.fill = PatternFill(start_color=name_colors[name],
                                               end_color=name_colors[name],
                                               fill_type="solid")
                
                # Team 2 players (row 2)
                for pi in range(2):
                    name = get_name(m, 1, pi)
                    cell = ws.cell(row=data_row2, column=col_start + 1 + pi)
                    cell.value = name
                    cell.font = Font(bold=True, size=10)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if name in name_colors:
                        cell.fill = PatternFill(start_color=name_colors[name],
                                               end_color=name_colors[name],
                                               fill_type="solid")
        
        current_row += 2
    
    # Set row heights for data rows
    for r in range(3, current_row):
        ws.row_dimensions[r].height = 22
    
    # Remove any remaining old data rows (from template) if we have fewer rounds
    for r in range(current_row, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).value = None
    
    # Save
    wb.save(output_path)
    return output_path


if __name__ == '__main__':
    if len(sys.argv) >= 3:
        matches_file = sys.argv[1]
        output_file = sys.argv[2]
        
        with open(matches_file, 'r', encoding='utf-8') as f:
            matches = json.load(f)
        
        male_names = None
        female_names = None
        date_str = "6月25日"
        venue_time = "19:00-22:00"
        duration = 30
        
        if '--date' in sys.argv:
            date_str = sys.argv[sys.argv.index('--date') + 1]
        if '--time' in sys.argv:
            venue_time = sys.argv[sys.argv.index('--time') + 1]
        if '--duration' in sys.argv:
            duration = int(sys.argv[sys.argv.index('--duration') + 1])
        if '--male' in sys.argv:
            with open(sys.argv[sys.argv.index('--male') + 1], 'r', encoding='utf-8') as f:
                male_names = [l.strip() for l in f if l.strip()]
        if '--female' in sys.argv:
            with open(sys.argv[sys.argv.index('--female') + 1], 'r', encoding='utf-8') as f:
                female_names = [l.strip() for l in f if l.strip()]
        
        result = export(matches, output_file, date_str, venue_time, duration,
                       male_names, female_names)
        print(f"Exported to {result}")
    else:
        print("Usage: python export_excel.py matches.json output.xlsx [--date 6月25日] [--time 19:00-22:00] [--duration 30] [--male males.txt] [--female females.txt]")
